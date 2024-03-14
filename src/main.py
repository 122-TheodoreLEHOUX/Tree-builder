"""Python program to create a file explorer with Tkinter UI,

This script is a mini app creating a folder structure visualisation using an excel file.
The script take into input the directory to scan and the destination of the excel containing the representation who will be created.
The user actions are : choose the directory to scan, choose the excel destination, click on launch button to execute script, close the window via the close button.
"""

# from zeutek_python import example_module
# from common import submodule

# import all components
# from the tkinter library
import json
import os
import tkinter
from tkinter import *
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Font,Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageTk
import customtkinter


CONFIG_FILE_PATH = "config/config.json"
ICON = "assets/zeutek-logo.ico"
COLORS = [
    '6D23FF',    # Purple
    'FFFFCCCC',  # Pale Pink
    'FFFFDDDD',  # Pale Rose
    'FFCCCCFF',  # Lavender
    'FFDDDDFF',  # Pale Lilac
    'FFCCFFCC',  # Pale Mint
    'FFE6E6E6',  # Light Gray
    'FFF0F0F0',  # Pale Silver
    'FFFFFF',    # White
    'FFFAFAFA',  # Snow
    '121212',    # Black
    '363636',    # Dark Grey
    'FFEAEAEA',  # Pale Ivory
    'FFF5F5F5'   # White Smoke
]
directory_size = 0 
index = 1

# Initialize TKinter
customtkinter.set_appearance_mode("dark")
window = customtkinter.CTk()
window.title('Directory scanner')
window.geometry("600x300")
window.iconbitmap(ICON)
window.resizable(width=False, height=False)

# Initialize Excel workbook
workbook = Workbook()
sheet = workbook.active
sheet.sheet_view.showGridLines = False
# Define a thick border style
THICK_BORDER = Border(top=Side(border_style="thick"),
                        right=Side(border_style="thick"),
                        bottom=Side(border_style="thick"),
                        left=Side(border_style="thick"))

# comment function
def rec_walk(dir, level):
    global index,directory_size, sheet

    sheet['A1'].font = Font(bold=True)
    sheet['A1'].border = THICK_BORDER
    sheet.cell(row=1, column=1, value=dir) # name of the dir scanned

    contents = os.listdir(dir)  # read the contents of dir
    for item in contents:  # loop over those contents
        index += 1
        if os.path.isdir(os.path.join(dir, item)):
            sheet.cell(row=index, column=level, value=item).fill = PatternFill(
                start_color=COLORS[level-1], end_color=COLORS[level-1], fill_type='solid')
            # recurse on subdirectories
            rec_walk(os.path.join(dir, item), level + 1)

        else:
            fp = os.path.join(dir, item)
            directory_size += os.path.getsize(fp)
            sheet.cell(row=index, column=level, value=item).fill = PatternFill(
                start_color=COLORS[level-1], end_color=COLORS[level-1], fill_type='solid')
            
    sheet.cell(row=1, column=2, value="Size : "+ str(directory_size/1000000) +" Mo") # display of the size of the folder
    sheet['B1'].font = Font(italic=True)
    sheet['B1'].border = THICK_BORDER
    return 0


# Browse function to use the browser explorer to choose a directory to scan
def browseFiles(var, label):
    filename = filedialog.askdirectory(
        initialdir="/", title="Select a Directory")
    var.set(filename)
    label.configure(text=filename)
    print(var.get())


def launch_function(workb, f_path, x_path,label):
    rec_walk(f_path.get(), 2) # run the recusive function on directory
    for column in sheet.columns:  #adjusting column widthness
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length+2
        sheet.column_dimensions[column_letter].width = adjusted_width

    complete_name = str(f_path) + "_structure.xlsx"
    xlsx_path = os.path.join(x_path.get(),complete_name)
    workb.save(xlsx_path) # save the file
    os.startfile(xlsx_path)# open it for the user
    label.configure(text="Done") # display done on window
    


# Program main loop
def main():
    # Read the config file
    with open(CONFIG_FILE_PATH) as json_file:
        config = json.load(json_file)

    # Initialize variables
    folder_path = StringVar()
    xslx_path = StringVar()

    # Configure the grid of the window
    window.grid_columnconfigure(0, weight=2)
    window.grid_columnconfigure(1, weight=1)
    window.grid_columnconfigure(2, weight=2,minsize=100)
    window.grid_rowconfigure(0, weight=1)
    window.grid_rowconfigure(1, weight=1)
    window.grid_rowconfigure(2, weight=1)
    window.grid_rowconfigure(3, weight=2)

    # Load and resize the image
    background_image= Image.open("assets/background.png")  # background image file
    width, height = background_image.size
    max_height = 300 
    resized_height = min(height, max_height)
    resized_width = int((resized_height / height) * width)
    resized_image = background_image.resize((resized_width, resized_height), Image.LANCZOS)
    tk_image = ImageTk.PhotoImage(resized_image) # Convert the image to Tkinter-compatible format
    frame = tkinter.Frame(window, bg="black")
    image_label = tkinter.Label(frame, image=tk_image,bg="black") # Create the label and place the image on it
    image_label.pack()

    # Add logo Icon
    logo = Image.open(ICON)  
    resized_logo = logo.resize((30, 30))
    tk_logo = ImageTk.PhotoImage(resized_logo)
    frame_logo = tkinter.Frame(window)
    logo_label = tkinter.Label(frame_logo, image=tk_logo,bg="black")
    logo_label.pack()

    # Create labels and buttons
    title = customtkinter.CTkLabel(master=window, text=" Zeutek Tree visualizer ", 
                                   font=("Segoe UI", 16,"bold"), fg_color=("transparent"))
    label_explore = customtkinter.CTkLabel(master=window, text="... ", wraplength=100, 
                                           font=("Segoe UI",12),fg_color=("transparent"))
    label_export = customtkinter.CTkLabel(master=window, text=os.path.dirname(os.path.abspath(__file__)), wraplength=100, 
                                          font=("Segoe UI",12), fg_color=("transparent"))
    label_finish = customtkinter.CTkLabel(master=window, text=" ", 
                                          font=("Segoe UI",12,"bold"), text_color=("olive"), fg_color=("transparent"))
    button_explore = customtkinter.CTkButton(master=window, text="Directory to scan",font=("Segoe UI", 13,"bold"),
                                             command=lambda: browseFiles(folder_path, label_explore))
    button_export = customtkinter.CTkButton(master=window, text="Export destination",font=("Segoe UI", 13,"bold"),
                                            command=lambda: browseFiles(xslx_path, label_export))
    button_launch = customtkinter.CTkButton(master=window, text="Launch",fg_color="purple",font=("Segoe UI", 14,"bold"),
                                            command=lambda: launch_function(workbook, folder_path, xslx_path,label_finish))


    # Place the labels and buttons in the grid
    frame.grid(row=0, column=0,rowspan=4, sticky="w")
    frame_logo.grid(row=0, column=1, columnspan=2,sticky="nw",pady=5)
    title.grid(row=0,column=1,sticky="n", columnspan=2,pady=8,padx=15)
    label_explore.grid(row=1, column=2, sticky="w")
    label_export.grid(row=2, column=2,sticky="w")
    label_finish.grid(row=3, column=2,sticky="w",pady=5)
    button_explore.grid(row=1, column=1, sticky="w")
    button_export.grid(row=2, column=1, sticky="w")
    button_launch.grid(row=3, column=1, sticky="w",pady=5)
    
    
    # Start the main loop
    window.mainloop()


if __name__ == "__main__":
    main()
